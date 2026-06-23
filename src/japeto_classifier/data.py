from __future__ import annotations

import hashlib
import json
from collections import Counter
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from japeto_classifier.constants import CATEGORY_ALIASES, CATEGORY_MAPPING, ORIGINAL_LABELS
from japeto_classifier.text import normalize_natural_text


REQUIRED_COLUMNS = {"user_message", "chatbot_response", "categories"}
OPTIONAL_COLUMNS = {
    "session_id",
    "message_time",
    "response_source",
    "intent_name",
    "category_grouped",
    "combined_text",
    "processed_text",
}


def _normalized_header(value: object) -> str:
    return str(value or "").strip().lower().replace(" ", "_")


def workbook_schema(path: Path) -> dict[str, object]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        raw_headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        headers = [_normalized_header(value) for value in raw_headers]
        return {
            "sheet": sheet.title,
            "rows_including_header": sheet.max_row,
            "columns": sheet.max_column,
            "headers": headers,
        }
    finally:
        workbook.close()


def read_source_records(path: Path) -> pd.DataFrame:
    """Read source records from the first worksheet."""
    return pd.read_excel(path, sheet_name=0)


def canonicalize_category(value: object) -> str:
    text = normalize_natural_text(value)
    if not text:
        return ""
    alias = CATEGORY_ALIASES.get(text.casefold())
    if alias:
        return alias
    lookup = {label.casefold(): label for label in ORIGINAL_LABELS}
    return lookup.get(text.casefold(), text)


def _stable_record_ids(frame: pd.DataFrame) -> list[str]:
    occurrences: Counter[str] = Counter()
    result: list[str] = []
    fields = ["session_id", "message_time", "user_message", "chatbot_response", "categories"]
    for row in frame.reindex(columns=fields, fill_value="").itertuples(index=False, name=None):
        payload = json.dumps([str(value or "") for value in row], ensure_ascii=False)
        digest = hashlib.sha256(payload.encode("utf-8")).hexdigest()[:20]
        occurrence = occurrences[digest]
        occurrences[digest] += 1
        result.append(digest if occurrence == 0 else f"{digest}-{occurrence}")
    return result


def normalize_records(frame: pd.DataFrame) -> pd.DataFrame:
    frame = frame.rename(columns={column: _normalized_header(column) for column in frame.columns}).copy()
    missing = sorted(REQUIRED_COLUMNS - set(frame.columns))
    if missing:
        raise ValueError(f"Dataset is missing required columns: {', '.join(missing)}")

    for column in REQUIRED_COLUMNS | OPTIONAL_COLUMNS:
        if column not in frame.columns:
            frame[column] = ""
    for column in ["session_id", "user_message", "chatbot_response", "categories"]:
        frame[column] = frame[column].map(normalize_natural_text)

    frame = frame.loc[(frame["chatbot_response"] != "") & (frame["categories"] != "")].copy()
    frame["categories"] = frame["categories"].map(canonicalize_category)
    unexpected = sorted(set(frame["categories"]) - set(ORIGINAL_LABELS))
    if unexpected:
        raise ValueError(f"Unmapped categories found: {', '.join(unexpected)}")
    frame["category_grouped"] = frame["categories"].map(CATEGORY_MAPPING)
    if frame["category_grouped"].isna().any():
        raise ValueError("Every original category must map to a grouped category")

    frame["record_id"] = _stable_record_ids(frame)
    missing_session = frame["session_id"] == ""
    frame.loc[missing_session, "session_id"] = frame.loc[missing_session, "record_id"]

    output_columns = [
        "record_id",
        "session_id",
        "message_time",
        "user_message",
        "chatbot_response",
        "response_source",
        "categories",
        "category_grouped",
        "intent_name",
    ]
    output = frame[output_columns].reset_index(drop=True)
    if not output["record_id"].is_unique:
        raise ValueError("Generated record IDs are not unique")
    return output


def file_sha256(path: Path, chunk_size: int = 1024 * 1024) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(chunk_size), b""):
            digest.update(chunk)
    return digest.hexdigest()


def ingest_dataset(source: Path, destination: Path) -> dict[str, object]:
    if not source.exists():
        raise FileNotFoundError(f"Dataset not found: {source}")
    destination.parent.mkdir(parents=True, exist_ok=True)
    normalized = normalize_records(read_source_records(source))
    normalized.to_parquet(destination, index=False)
    summary = {
        "source": str(source),
        "destination": str(destination),
        "source_sha256": file_sha256(source),
        "rows": len(normalized),
        "sessions": int(normalized["session_id"].nunique()),
        "original_labels": int(normalized["categories"].nunique()),
        "grouped_labels": int(normalized["category_grouped"].nunique()),
        "context_eligible_rows": int((normalized["user_message"] != "").sum()),
    }
    (destination.parent / "records.manifest.json").write_text(
        json.dumps(summary, indent=2), encoding="utf-8"
    )
    return summary


def processed_records_are_current(source: Path, destination: Path) -> bool:
    manifest_path = destination.parent / "records.manifest.json"
    if not source.exists() or not destination.exists() or not manifest_path.exists():
        return False
    try:
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return False
    return manifest.get("source_sha256") == file_sha256(source)
